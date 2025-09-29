import os, re, io, csv
from flask import Flask, render_template, request, redirect, url_for, flash
from sqlalchemy import create_engine, Column, Integer, String, Numeric, Text, ForeignKey
from sqlalchemy.orm import sessionmaker, declarative_base, relationship

# --- App ---
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "change-this-secret")

# Auth for uploads/settings
UPLOAD_PASSWORD = os.environ.get("UPLOAD_PASSWORD", "Ionut2025")

# Database (Render's DATABASE_URL) - fallback to SQLite for local testing

db_url = os.environ.get("DATABASE_URL", "").replace("postgres://", "postgresql+psycopg://")
# Normalize to psycopg v3 driver if plain postgresql:// is provided
if db_url.startswith("postgresql://"):
    db_url = db_url.replace("postgresql://", "postgresql+psycopg://", 1)

connect_args = {"sslmode": "require"}  # Render Postgres usually requires SSL

engine = create_engine(
    db_url,
    pool_pre_ping=True,
    pool_recycle=1800,  # recycle stale connections
    pool_size=5,
    max_overflow=10,
    connect_args=connect_args
)
SessionLocal = sessionmaker(bind=engine)

Base = declarative_base()

# --- Models ---
class Setting(Base):
    __tablename__ = "settings"
    id = Column(Integer, primary_key=True)
    key = Column(String(100), unique=True, nullable=False)
    value = Column(Text, nullable=True)

class Category(Base):
    __tablename__ = "categories"
    id = Column(Integer, primary_key=True)
    slug = Column(String(200), unique=True, nullable=False)
    name = Column(String(300), nullable=False)

class Ranking(Base):
    __tablename__ = "rankings"
    id = Column(Integer, primary_key=True)
    category_id = Column(Integer, ForeignKey("categories.id"), nullable=False)
    position = Column(String(50))
    competitor = Column(String(200))
    club = Column(String(200))
    execution = Column(Numeric(8,3), nullable=True)
    artistry = Column(Numeric(8,3), nullable=True)
    difficulty = Column(Numeric(8,3), nullable=True)
    line_penalty = Column(Numeric(8,3), nullable=True)
    chair_penalty = Column(Numeric(8,3), nullable=True)
    diff_penalty = Column(Numeric(8,3), nullable=True)
    total = Column(Numeric(8,3), nullable=True)
    category = relationship("Category")

Base.metadata.create_all(engine)

# --- Utils ---
def slugify(text:str)->str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")

def to_num(v):
    try:
        if v is None or v == "":
            return None
        return float(str(v).replace(",", "."))
    except:
        return None

def get_settings():
    s = SessionLocal()
    out = {}
    try:
        for row in s.query(Setting).all():
            out[row.key] = row.value
    finally:
        s.close()
    out.setdefault("title", "National Gymnastics Rankings – Romania")
    out.setdefault("subtitle", "Official RENC")
    out.setdefault("event_date", "")
    out.setdefault("location", "")
    return out

def set_setting(key, value):
    s = SessionLocal()
    try:
        rec = s.query(Setting).filter_by(key=key).first()
        if rec:
            rec.value = value
        else:
            s.add(Setting(key=key, value=value))
        s.commit()
    finally:
        s.close()

# --- Seed default categories & demo rows (only if DB new) ---
DEFAULT_CATS = [
    "INDIVIDUAL WOMEN - YOUTH TRIC - KIDS DEVELOPMENT",
    "INDIVIDUAL MEN - YOUTH",
    "MIXED PAIR - NATIONAL DEVELOPMENT",
    "INDIVIDUAL WOMEN - KIDS DEVELOPMENT",
    "MIXED PAIR - JUNIORS",
    "TRIO - NATIONAL DEVELOPMENT",
    "INDIVIDUAL WOMEN - JUNIORS",
    "INDIVIDUAL MEN - JUNIORS",
    "INDIVIDUAL WOMEN - NATIONAL DEVELOPMENT",
    "INDIVIDUAL MEN - KIDS DEVELOPMENT",
    "MIXED PAIR - YOUTH",
    "INDIVIDUAL MEN - NATIONAL DEVELOPMENT",
    "GROUP - YOUTH",
    "GROUP - JUNIORS",
    "INDIVIDUAL MEN - SENIORS",
    "TRIO - YOUTH",
    "TRIO - JUNIORS",
    "GROUP - KIDS DEVELOPMENT",
    "GROUP - NATIONAL DEVELOPMENT",
    "AEROBIC DANCE - YOUTH",
    "AEROBIC DANCE - JUNIORS"
]

REQUIRED_HEADERS = [
    "POSITION","COMPETITOR","CLUB","EXECUTION","ARTISTRY",
    "DIFFICULTY","LINE PENALTY","CHAIR PENALTY","DIFF PENALTY","TOTAL"
]


def seed_if_empty():
    s = SessionLocal()
    try:
        if s.query(Category).count() == 0:
            for name in DEFAULT_CATS:
                s.add(Category(slug=slugify(name), name=name))
            s.commit()
        if s.query(Ranking).count() == 0:
            cm = s.query(Category).filter_by(slug=slugify("Individual Men – Senior")).first()
            cw = s.query(Category).filter_by(slug=slugify("Individual Women – Senior")).first()
            if cm:
                s.add(Ranking(category_id=cm.id, position="1", competitor="Ion Popescu", club="Steaua", total=15.200))
                s.add(Ranking(category_id=cm.id, position="2", competitor="Mihai Ionescu", club="Dinamo", total=14.850))
                s.add(Ranking(category_id=cm.id, position="3", competitor="Andrei Georgescu", club="CSM Cluj", total=14.300))
            if cw:
                s.add(Ranking(category_id=cw.id, position="1", competitor="Simona Ionescu", club="Dinamo", total=14.800))
            s.commit()
        if not s.query(Setting).filter_by(key="title").first():
            set_setting("title", "National Gymnastics Rankings – Romania")
            set_setting("subtitle", "Official RENC")
    finally:
        s.close()

seed_if_empty()

# --- Routes ---


@app.route("/", methods=["GET","HEAD"])
def home():
    if request.method == "HEAD":
        return ("", 200)
    selected = (request.args.get("category") or "").strip()
    s = SessionLocal()
    settings = get_settings()
    try:
        # Dropdown source (to be always available)
        all_categories = [
            type('Obj',(object,),{'slug':c.slug,'name':c.name})()
            for c in s.query(Category).order_by(Category.name).all()
        ]
        categories = {}
        if selected:
            c = s.query(Category).filter_by(slug=selected).first()
            if c:
                rows = (s.query(Ranking)
                          .filter_by(category_id=c.id)
                          .order_by(Ranking.total.desc().nullslast())
                          .all())
                categories[c.slug] = type("Obj",(object,),{"slug":c.slug,"name":c.name,"rows":rows})()
        else:
            for c in s.query(Category).order_by(Category.name).all():
                rows = (s.query(Ranking)
                          .filter_by(category_id=c.id)
                          .order_by(Ranking.total.desc().nullslast())
                          .all())
                categories[c.slug] = type("Obj",(object,),{"slug":c.slug,"name":c.name,"rows":rows})()
    finally:
        s.close()
    return render_template("home.html", settings=settings, categories=categories, selected=selected, all_categories=all_categories)

@app.route("/category/<slug>")
def category_page(slug):
    s = SessionLocal()
    settings = get_settings()
    try:
        cat = s.query(Category).filter_by(slug=slug).first()
        rows = s.query(Ranking).filter_by(category_id=cat.id).order_by(Ranking.id).all() if cat else []
    finally:
        s.close()
    return render_template("category.html", settings=settings, cat=cat, rows=rows)


@app.route("/upload", methods=["GET","POST"])
def upload():
    settings = get_settings()
    if request.method == "POST":
        pwd = request.form.get("password","").strip()
        if pwd != UPLOAD_PASSWORD:
            flash("Incorrect password.", "error")
            return redirect(url_for("upload"))
        # category resolution
        cat_slug = request.form.get("category_slug","").strip()
        cat_name = request.form.get("category_name","").strip()
        if not cat_slug and not cat_name:
            flash("Please select or enter a category.", "error")
            return redirect(url_for("upload"))
        s = SessionLocal()
        try:
            if cat_slug:
                cat = s.query(Category).filter_by(slug=cat_slug).first()
                if not cat:
                    flash("Category not found.", "error")
                    return redirect(url_for("upload"))
            else:
                slug = slugify(cat_name)
                cat = s.query(Category).filter_by(slug=slug).first()
                if not cat:
                    cat = Category(slug=slug, name=cat_name)
                    s.add(cat); s.commit(); s.refresh(cat)

            file = request.files.get("file")
            if not file or file.filename == "":
                flash("Please choose a CSV or XLSX file.", "error")
                return redirect(url_for("upload"))

            name = file.filename.lower()
            count = 0
            if name.endswith(".csv"):
                data = io.StringIO(file.stream.read().decode("utf-8"))
                reader = csv.DictReader(data)
                fields = [ (h or '').strip() for h in (reader.fieldnames or []) ]
                missing = [h for h in REQUIRED_HEADERS if h not in fields]
                if missing:
                    flash("Missing required columns: " + ", ".join(missing), "error")
                    return redirect(url_for("category_page", slug=cat.slug))
                for r in reader:
                    s.add(Ranking(
                        category_id=cat.id,
                        position=str(r.get("POSITION") or ""),
                        competitor=str(r.get("COMPETITOR") or ""),
                        club=str(r.get("CLUB") or ""),
                        execution=to_num(r.get("EXECUTION")),
                        artistry=to_num(r.get("ARTISTRY")),
                        difficulty=to_num(r.get("DIFFICULTY")),
                        line_penalty=to_num(r.get("LINE PENALTY")),
                        chair_penalty=to_num(r.get("CHAIR PENALTY")),
                        diff_penalty=to_num(r.get("DIFF PENALTY")),
                        total=to_num(r.get("TOTAL"))
                    )); count += 1
            elif name.endswith(".xlsx"):
                import openpyxl
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                headers = [str((c.value or "")).strip() for c in ws[1]]
                missing = [h for h in REQUIRED_HEADERS if h not in headers]
                if missing:
                    flash("Missing required columns: " + ", ".join(missing), "error")
                    return redirect(url_for("category_page", slug=cat.slug))
                idx = {h:i for i,h in enumerate(headers)}
                def g(row, key):
                    return row[idx[key]].value if key in idx else None
                for row in ws.iter_rows(min_row=2, values_only=False):
                    s.add(Ranking(
                        category_id=cat.id,
                        position=str(g(row,"POSITION") or ""),
                        competitor=str(g(row,"COMPETITOR") or ""),
                        club=str(g(row,"CLUB") or ""),
                        execution=to_num(g(row,"EXECUTION")),
                        artistry=to_num(g(row,"ARTISTRY")),
                        difficulty=to_num(g(row,"DIFFICULTY")),
                        line_penalty=to_num(g(row,"LINE PENALTY")),
                        chair_penalty=to_num(g(row,"CHAIR PENALTY")),
                        diff_penalty=to_num(g(row,"DIFF PENALTY")),
                        total=to_num(g(row,"TOTAL"))
                    )); count += 1
            else:
                flash("Unsupported file type. Use CSV or XLSX.", "error")
                return redirect(url_for("category_page", slug=cat.slug))

            s.commit()
            flash(f"Uploaded {{count}} rows into '{{cat.name}}'.", "success")
            return redirect(url_for("category_page", slug=cat.slug))
        finally:
            s.close()
    return render_template("upload.html", settings=settings)



@app.route("/settings", methods=["GET","POST"])
def settings_page():
    settings = get_settings()
    s = SessionLocal()
    try:
        cats = s.query(Category).order_by(Category.name).all()
        # compute counts per category
        counts = {}
        for c in cats:
            counts[c.slug] = s.query(Ranking).filter_by(category_id=c.id).count()
    finally:
        s.close()

    if request.method == "POST":
        pwd = request.form.get("password","").strip()
        if pwd != UPLOAD_PASSWORD:
            flash("Incorrect password for settings.", "error")
            return redirect(url_for("settings_page"))

        action = request.form.get("action","save_settings")
        if action == "save_settings":
            set_setting("title", request.form.get("title","").strip())
            set_setting("subtitle", request.form.get("subtitle","").strip())
            set_setting("event_date", request.form.get("event_date","").strip())
            set_setting("location", request.form.get("location","").strip())
            flash("Settings saved.", "success")

        elif action == "add_cat":
            name = request.form.get("new_category","").strip()
            if name:
                s = SessionLocal()
                try:
                    slug = slugify(name)
                    if not s.query(Category).filter_by(slug=slug).first():
                        s.add(Category(slug=slug, name=name)); s.commit()
                        flash(f"Added category '{name}'.", "success")
                    else:
                        flash("Category already exists.", "error")
                finally:
                    s.close()
        return redirect(url_for("settings_page"))
    return render_template("settings.html", settings=settings, cats=cats, cat_counts=counts)

@app.route("/privacy")
def privacy():
    return render_template("privacy.html", settings=get_settings())

@app.route("/terms")
def terms():
    return render_template("terms.html", settings=get_settings())

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT","5000")), debug=True)


@app.before_request
def _allow_head():
    if request.method == "HEAD":
        return ("", 200)


@app.route("/category/<slug>/delete", methods=["POST"])
def delete_category_data(slug):
    pwd = request.form.get("password","").strip()
    if pwd != UPLOAD_PASSWORD:
        flash("Parola incorecta.", "error")
        return redirect(url_for("category_page", slug=slug))
    s = SessionLocal()
    try:
        cat = s.query(Category).filter_by(slug=slug).first()
        if not cat:
            flash("Categoria nu exista.", "error")
            return redirect(url_for("home"))
        deleted = s.query(Ranking).filter_by(category_id=cat.id).delete()
        s.commit()
        flash(f"Sters {deleted} randuri pentru categoria '{cat.name}'.", "success")
    finally:
        s.close()
    return redirect(url_for("category_page", slug=slug))


@app.route("/delete-by-name", methods=["POST"])
def delete_by_name():
    pwd = request.form.get("password","").strip()
    if pwd != UPLOAD_PASSWORD:
        flash("Parola incorecta.", "error")
        return redirect(url_for("upload"))
    name = request.form.get("category_name","").strip()
    if not name:
        flash("Completeaza numele categoriei.", "error")
        return redirect(url_for("upload"))
    slug = slugify(name)
    s = SessionLocal()
    try:
        cat = s.query(Category).filter_by(slug=slug).first()
        if not cat:
            flash("Categoria nu exista.", "error")
            return redirect(url_for("upload"))
        deleted = s.query(Ranking).filter_by(category_id=cat.id).delete()
        s.commit()
        flash(f"Sters {deleted} randuri pentru categoria '{cat.name}'.", "success")
    finally:
        s.close()
    return redirect(url_for("category_page", slug=slug))


@app.route("/category/<slug>/delete-category", methods=["POST"])
def delete_category(slug):
    pwd = request.form.get("password","").strip()
    if pwd != UPLOAD_PASSWORD:
        flash("Parola incorecta.", "error")
        return redirect(url_for("category_page", slug=slug))
    s = SessionLocal()
    try:
        cat = s.query(Category).filter_by(slug=slug).first()
        if not cat:
            flash("Categoria nu exista.", "error")
            return redirect(url_for("home"))
        # sterge randurile
        s.query(Ranking).filter_by(category_id=cat.id).delete()
        # sterge categoria
        name = cat.name
        s.delete(cat)
        s.commit()
        flash(f"Categoria '{name}' a fost stearsa complet.", "success")
    finally:
        s.close()
    return redirect(url_for("home"))


@app.route("/settings/category/delete", methods=["POST"])
def settings_delete_category():
    pwd = request.form.get("password","").strip()
    if pwd != UPLOAD_PASSWORD:
        flash("Parola incorecta.", "error")
        return redirect(url_for("settings_page"))
    slug = request.form.get("slug","").strip()
    s = SessionLocal()
    try:
        cat = s.query(Category).filter_by(slug=slug).first()
        if not cat:
            flash("Categoria nu exista.", "error")
            return redirect(url_for("settings_page"))
        s.query(Ranking).filter_by(category_id=cat.id).delete()
        name = cat.name
        s.delete(cat); s.commit()
        flash(f"Categoria '{name}' a fost stearsa complet.", "success")
    finally:
        s.close()
    return redirect(url_for("settings_page"))

@app.route("/settings/category/rename", methods=["POST"])
def settings_rename_category():
    pwd = request.form.get("password","").strip()
    if pwd != UPLOAD_PASSWORD:
        flash("Parola incorecta.", "error")
        return redirect(url_for("settings_page"))
    slug = request.form.get("slug","").strip()
    new_name = request.form.get("new_name","").strip()
    if not new_name:
        flash("Introdu un nume nou.", "error")
        return redirect(url_for("settings_page"))
    s = SessionLocal()
    try:
        cat = s.query(Category).filter_by(slug=slug).first()
        if not cat:
            flash("Categoria nu exista.", "error")
            return redirect(url_for("settings_page"))
        new_slug = slugify(new_name)
        # check duplicate
        if s.query(Category).filter(Category.id != cat.id, Category.slug == new_slug).first():
            flash("Exista deja o categorie cu acest nume.", "error")
            return redirect(url_for("settings_page"))
        cat.name = new_name
        cat.slug = new_slug
        s.commit()
        flash("Categoria a fost redenumita.", "success")
    finally:
        s.close()
    return redirect(url_for("settings_page"))


# --- Simple i18n ---
from flask import make_response, g

I18N = {
    "en": {
        "Upload": "Upload",
        "Settings": "Settings",
        "Privacy": "Privacy",
        "Terms": "Terms",
        "Date": "Date",
        "Location": "Location",
        "Upload to this category": "Upload to this category",
        "Password": "Password",
        "File (CSV or XLSX)": "File (CSV or XLSX)",
        "Headers required": "Headers required",
        "Delete file (all rows)": "Delete file (all rows)",
        "Delete CATEGORY (permanent)": "Delete CATEGORY (permanent)",
    },
    "ro": {
        "Upload": "Incarca",
        "Settings": "Setari",
        "Privacy": "Confidentialitate",
        "Terms": "Termeni",
        "Date": "Data",
        "Location": "Locatie",
        "Upload to this category": "Incarca in aceasta categorie",
        "Password": "Parola",
        "File (CSV or XLSX)": "Fisier (CSV sau XLSX)",
        "Headers required": "Header-e necesare",
        "Delete file (all rows)": "Sterge fisierul (toate randurile)",
        "Delete CATEGORY (permanent)": "Sterge CATEGORIA (definitiv)",
    },
}

def get_lang():
    code = request.cookies.get("lang", "en")
    return "ro" if code == "ro" else "en"

@app.context_processor
def inject_i18n():
    code = get_lang()
    def t(key):
        return I18N.get(code, I18N["en"]).get(key, key)
    return {"t": t, "lang_code": code}

@app.route("/set-lang/<code>")
def set_lang(code):
    code = "ro" if code == "ro" else "en"
    nxt = request.args.get("next") or request.referrer or url_for("home")
    resp = make_response(redirect(nxt))
    resp.set_cookie("lang", code, max_age=60*60*24*365)  # 1 an
    return resp
